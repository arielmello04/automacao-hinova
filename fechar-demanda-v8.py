from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
from datetime import datetime

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/FECHAR DEMANDA.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def wait_until_element_clickable(by, value, timeout=10):
    """Espera até que o elemento especificado esteja clicável."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )

def open_page_and_accept():
    try:
        print("Abrindo a página de agendamento...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]'))
        )
        print("Página de agendamento carregada.")

        print("Clicando no botão 'Aceitar e fechar'...")
        accept_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]')
        accept_button.click()
        print("Botão 'Aceitar e fechar' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e clicar no aviso:", e)
        sheet['D1'] = f"Erro ao abrir a página: {e}" 
        workbook.save(file_path)
        driver.quit()
        exit()

def login():
    try:
        print("Preenchendo o formulário de login...")
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]'))
        )
        client_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]')
        client_field.clear()
        client_field.send_keys("")

        name_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.nome"]')
        name_field.clear()
        name_field.send_keys("")

        password_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.senha"]')
        password_field.clear()
        password_field.send_keys("")

        print("Clicando no botão 'Entrar'...")
        login_button = driver.find_element(By.CSS_SELECTOR, 'button[ng-click="login()"]')
        login_button.click()
        print("Botão 'Entrar' clicado com sucesso.")

        time.sleep(5)

    except Exception as e:
        print("Ocorreu um erro ao preencher o formulário de login e clicar em 'Entrar':", e)
        sheet['D1'] = f"Erro ao fazer login: {e}" 
        workbook.save(file_path)
        driver.quit()
        exit()

def check_and_fill_bairro(sheet, row):
    """Verifica se o campo 'Bairro' está vazio e preenche com 'CENTRO' se necessário."""
    try:
        bairro_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="bairro"][ng-model="formData.bairro"]'))
        )
        
        bairro_value = bairro_input.get_attribute('value').strip()
        if not bairro_value:
            bairro_input.clear()
            bairro_input.send_keys("CENTRO")
            sheet[f'G{row}'] = "Bairro preenchido"
            print("Campo 'Bairro' estava vazio, preenchido com 'CENTRO'.")
        else:
            print("Campo 'Bairro' já está preenchido, seguindo com o processo.")
    
    except Exception as e:
        print(f"Erro ao verificar ou preencher o campo 'Bairro': {e}")
        sheet[f'G{row}'] = f"Erro ao preencher Bairro: {e}"

def search_plates(start_row):
    try:
        print("Navegando para a página de busca de placas...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="search"]'))
        )
        print("Página de busca carregada.")

        for row_number in range(start_row, sheet.max_row + 1):
            plate = sheet[f'B{row_number}'].value
            technician_name = sheet[f'C{row_number}'].value
            if not plate or not technician_name:
                print(f"Nenhuma placa ou técnico encontrado na linha {row_number}. Encerrando.")
                break
            
            print(f"Buscando a placa: {plate}")
            
            search_field = driver.find_element(By.CSS_SELECTOR, 'input[type="search"]')
            search_field.clear()
            search_field.send_keys(plate)
            
            time.sleep(4) 
            tr_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'tbody tr'))
            )
            try:
                rows = driver.find_elements(By.CSS_SELECTOR, 'tbody tr')
                found = False
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    plate_in_row = cells[3].text.strip()
                    installation_status = cells[8].text.strip()
                    
                    if plate_in_row == plate and installation_status == 'INSTALAÇÃO':
                        print(f"Placa {plate} encontrada com status 'INSTALAÇÃO'.")
                        edit_button = row.find_element(By.CSS_SELECTOR, 'a.btn-info')
                        edit_button.click()
                        print("Botão 'Editar' clicado com sucesso.")
                        
                        print("Alterando o status para 'DEMANDA CONCLUÍDA'...")
                        status_select = wait_until_element_clickable(By.CSS_SELECTOR, 'select[ng-model="formData.situacao_id"]')
                        status_select.click()
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'select[ng-model="formData.situacao_id"] option[value="string:1144"]'))
                        )
                        status_select.find_element(By.CSS_SELECTOR, 'option[value="string:1144"]').click()

                        today_date = datetime.today().strftime('%d/%m/%Y')
                        print(f"Inserindo a data de hoje: {today_date}...")
                        date_input = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="formData.data_conclusao"]')
                        date_input.clear()
                        date_input.send_keys(today_date)

                        print("Aguardando 4 segundos antes de interagir com o select de técnicos...")
                        time.sleep(4)

                        try:
                            close_button = driver.find_element(By.CSS_SELECTOR, 'div.toast button.close')
                            close_button.click()
                        except:
                            print("Não foi possível encontrar o botão de fechar a notificação.")

                        print(f"Esperando o select de técnicos estar disponível...")
                        technician_select = wait_until_element_clickable(By.CSS_SELECTOR, 'select[ng-model="formData.tecnico_id"]')
                        technician_select.click()

                        print(f"Buscando técnico: {technician_name}...")
                        WebDriverWait(driver, 20).until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'select[ng-model="formData.tecnico_id"] option'))
                        )
                        options = technician_select.find_elements(By.TAG_NAME, 'option')
                        
                        if technician_name.upper() == "IAGO":
                            print(f"Selecionando técnico específico: IAGO DANIEL LIMA OLIVEIRA...")
                            driver.execute_script("arguments[0].click();", technician_select.find_element(By.CSS_SELECTOR, 'option[value="string:67"]'))
                        else:
                            found = False
                            for option in options:
                                if technician_name.upper() in option.text.upper():
                                    print(f"Selecionando técnico: {option.text}...")
                                    option.click()
                                    found = True
                                    break

                            if not found:
                                print(f"Técnico '{technician_name}' não encontrado.")
                                sheet[f'D{row_number}'] = "Técnico não encontrado"
                                break

                        check_and_fill_bairro(sheet, row_number)

                        print("Clicando no botão 'Gravar'...")
                        save_button = wait_until_element_clickable(By.CSS_SELECTOR, 'button[ng-disabled="processandoEnvio"]')
                        save_button.click()
                        print("Botão 'Gravar' clicado com sucesso.")
                        sheet[f'E{row_number}'] = "Agendamento fechado com sucesso."
                        print("Esperando o modal de confirmação aparecer...")
                        ok_button = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.modal-footer.ng-scope button.btn.btn-success[ng-click="ok()"]'))
                        )
                        print("Clicando no botão 'OK' do modal de confirmação...")
                        ok_button.click()
                        print("Botão 'OK' clicado com sucesso.")

                        print("Voltando para a página de busca...")
                        driver.get('site')
                        WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="search"]'))
                        )

                        found = True
                        break  

                if not found:
                    print(f"Placa {plate} não encontrada ou status não é 'INSTALAÇÃO'.")
                    sheet[f'D{row_number}'] = "Placa não encontrada ou status incorreto"
                    continue 

            except Exception as e:
                print(f"Não foi possível processar a placa {plate}:", e)
                sheet[f'D{row_number}'] = f"Falha ao processar: {e}" 
                continue 

    except Exception as e:
        print(f"Ocorreu um erro ao buscar placas:", e)
        sheet['D1'] = f"Erro na busca de placas: {e}"  
        workbook.save(file_path)
        driver.quit()
        exit()

def main():
    open_page_and_accept()
    login()

    try:
        start_row = int(input("Digite a linha inicial para começar a captura das placas: "))
    except ValueError:
        print("Valor inválido para a linha inicial. Encerrando.")
        driver.quit()
        exit()

    search_plates(start_row)

    print("Salvando o arquivo Excel...")
    workbook.save(file_path)
    print("Arquivo Excel salvo com sucesso.")

    input("Pressione Enter para encerrar a execução e fechar o navegador...")
    driver.quit()

if __name__ == "__main__":
    main()
