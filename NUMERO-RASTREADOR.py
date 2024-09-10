from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import re
import sys

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/NUM RASTREADOR.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

chrome_options = Options()
chrome_options.add_argument("--mute-audio") 

service = Service(driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

def close_overlays():
    try:
        print("Verificando se há overlays bloqueando a tela...")
        WebDriverWait(driver, 5).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, 'div.overlay'))
        )
        print("Overlays fechados ou não presentes.")
    except Exception as e:
        print("Nenhum overlay ou erro ao fechar:", e)

def open_page_and_login():
    try:
        print("Abrindo a página de login...")
        driver.get('site')

        print("Preenchendo o formulário de login...")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'login'))
        )
        login_field = driver.find_element(By.ID, 'login')
        login_field.send_keys("")

        password_field = driver.find_element(By.ID, 'senha')
        password_field.send_keys("")

        close_overlays()

        print("Clicando no botão 'Entrar'...")
        enter_button = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
        enter_button.click()
        print("Botão 'Entrar' clicado com sucesso.")
        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e fazer login:", e)
        driver.quit()
        sys.exit()

def search_plates(start_row):
    try:
        print("Navegando para a página de busca de veículos...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Filtrar"]'))
        )
        print("Página de busca de veículos carregada.")

        for row_number in range(start_row, sheet.max_row + 1):
            plate = sheet[f'B{row_number}'].value
            if not plate:
                print(f"Celula vazia encontrada na linha {row_number}. Encerrando.")
                break

            plate = str(plate).strip()
            print(f"Buscando a placa: {plate}")

            search_field = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Filtrar"]')
            search_field.clear()
            search_field.send_keys(plate)

            close_overlays()

            search_icon = driver.find_element(By.CSS_SELECTOR, 'glyph.datatable-search-icon')
            search_icon.click()

            time.sleep(2)

            found = False
            rows = driver.find_elements(By.CSS_SELECTOR, 'tbody tr')
            for row in rows:
                tds = row.find_elements(By.CSS_SELECTOR, 'td')
                for td in tds:
                    if plate == str(td.text).strip():
                        found = True
                        print("Placa encontrada na tabela.")
                        suspend_icon = row.find_element(By.CSS_SELECTOR, 'glyph.datatable-suspender-icon')
                        suspend_icon.click()

                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="button"][value="Sim"]'))
                        ).click()

                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="button"][value="Ok"]'))
                        ).click()

                        break
                if found:
                    break

            if not found:
                print(f"Placa {plate} não encontrada na tabela de veículos.")
                sheet[f'D{row_number}'] = "Placa não encontrada em veículos."
                workbook.save(file_path)

    except Exception as e:
        print(f"Ocorreu um erro ao buscar placas:", e)
        sheet[f'E{row_number}'] = "Erro ao buscar em veículos."
        workbook.save(file_path)
        driver.quit()
        sys.exit()

def search_equipment(start_row):
    try:
        print("Navegando para a página de busca de equipamentos...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Filtrar"]'))
        )
        print("Página de busca de equipamentos carregada.")

        select_element = driver.find_element(By.ID, 'bar_operation_select')
        if select_element.get_attribute('value') != 'disponiveis':
            print("A opção 'Disponíveis' não está selecionada. Ajustando...")
            select_element.click()
            select_option = driver.find_element(By.XPATH, "//select[@id='bar_operation_select']/option[@value='disponiveis']")
            select_option.click()
            time.sleep(2)

        for row_number in range(start_row, sheet.max_row + 1):
            number = sheet[f'B{row_number}'].value
            if not number:
                print(f"Celula vazia encontrada na linha {row_number}. Encerrando.")
                break

            number = str(number).strip()
            print(f"Buscando o número: {number}")

            search_field = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Filtrar"]')
            search_field.clear()
            search_field.send_keys(number)

            close_overlays()

            search_icon = driver.find_element(By.CSS_SELECTOR, 'glyph.datatable-search-icon')
            search_icon.click()

            time.sleep(2)

            found = False
            rows = driver.find_elements(By.CSS_SELECTOR, 'tbody tr')
            for row in rows:
                try:
                    equipment_cell = row.find_element(By.CSS_SELECTOR, 'td')
                    if equipment_cell and number == equipment_cell.text.strip():
                        found = True
                        print("Número encontrado na tabela.")

                        module_cell = row.find_element(By.CSS_SELECTOR, 'td.modulo a.data_inspect')
                        text = module_cell.text
                        number_found = re.sub(r'\D', '', text)
                        if number_found:
                            sheet[f'C{row_number}'] = number_found
                            print(f"Número extraído: {number_found}")

                        less_circle_icon = row.find_element(By.CSS_SELECTOR, 'glyph.datatable-circulo-menos-icon')
                        less_circle_icon.click()

                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="button"][value="Ok"].btn-modal-primary'))
                        ).click()

                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="button"][value="Ok"].btn-modal-primary'))
                        ).click()

                        break
                except Exception as e:
                    print(f"Erro ao verificar o número em equipamentos {number}: {e}")
                    sheet[f'D{row_number}'] = "Número não encontrado em equipamentos"
                    workbook.save(file_path)

            if not found:
                print(f"Número {number} não encontrado na tabela de equipamentos.")
                sheet[f'E{row_number}'] = "Número não encontrado em equipamentos"
                workbook.save(file_path)

    except Exception as e:
        print(f"Ocorreu um erro ao buscar números:", e)
        sheet[f'F{row_number}'] = "Erro ao buscar em equipamentos."
        workbook.save(file_path)
        driver.quit()
        sys.exit()

def main():
    try:
        open_page_and_login()

        start_row = int(input("Por favor, insira o número da linha inicial: "))

        search_plates(start_row)
        search_equipment(start_row)

    except Exception as e:
        print(f"Ocorreu um erro durante a execução: {e}")

    finally:
        try:
            print("Salvando o arquivo Excel...")
            workbook.save(file_path)
            print("Arquivo Excel salvo com sucesso.")
        except Exception as save_error:
            print(f"Erro ao salvar o arquivo Excel: {save_error}")

        input("Pressione Enter para encerrar a execução e fechar o navegador...")
        driver.quit()

if __name__ == "__main__":
    main()
