from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from datetime import datetime
import openpyxl

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/VINCULAR RASTREADOR.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def wait_for_angular():
    """Aguardar o AngularJS concluir suas atualizações."""
    try:
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return window.angular !== undefined && angular.element(document).injector().get("$http").pendingRequests.length === 0')
        )
        print("AngularJS atualizado.")
    except Exception as e:
        print("Erro ao aguardar o AngularJS:", e)

def wait_until_element_clickable(by, value, timeout=10):
    """Espera até que o elemento especificado esteja clicável."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )

def open_page_and_accept():
    try:
        print("Abrindo a página de vinculo de rastreador...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]'))
        )
        print("Página de agendamento carregada.")

        print("Clicando no botão 'Aceitar e fechar'...")
        accept_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]')
        accept_button.click()
        print("Botão 'Aceitar e fechar' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e clicar no aviso:", e)
        driver.quit()
        exit()

def login():
    try:
        print("Preenchendo o formulário de login...")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]'))
        )
        client_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]')
        client_field.clear()
        client_field.send_keys("")

        name_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.nome"]')
        name_field.clear()
        name_field.send_keys("")

        password_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.senha"]')
        password_field.clear()
        password_field.send_keys("")

        print("Clicando no botão 'Entrar'...")
        login_button = driver.find_element(By.CSS_SELECTOR, 'button[ng-click="login()"]')
        login_button.click()
        print("Botão 'Entrar' clicado com sucesso.")

        time.sleep(5)

    except Exception as e:
        print("Ocorreu um erro ao preencher o formulário de login e clicar em 'Entrar':", e)
        driver.quit()
        exit()

def navigate_to_vincular_rastreador_page():
    try:
        print("Navegando para a página de vinculação de rastreador...")
        driver.get('site')
        print("Página de vinculação de rastreador carregada.")
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'placa_vinculo'))
        )
        
    except Exception as e:
        print("Ocorreu um erro ao navegar para a página de vinculação de rastreador:", e)
        driver.quit()
        exit()

def reset_search_input(input_field, value):
    try:
        input_field.send_keys(Keys.BACKSPACE) 
        input_field.send_keys(value[-1])  
        print("Último caractere removido e reinserido.")
        return True
    except Exception as e:
        print(f"Erro ao tentar redefinir a entrada de busca: {e}")
        return False

def insert_placa_value(row):
    try:
        print(f"Inserindo valor da coluna B da linha {row}...")
        placa_value = sheet[f'B{row}'].value
        if not placa_value:
            print(f"Nenhum valor encontrado na coluna B da linha {row}. Continuando para a próxima linha.")
            sheet[f'G{row}'] = "Nenhum valor encontrado"
            workbook.save(file_path)
            return False

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'placa_vinculo'))
        )
        placa_field = driver.find_element(By.NAME, 'placa_vinculo')
        placa_field.clear()
        placa_field.send_keys(placa_value)
        print(f"Valor '{placa_value}' inserido no campo de texto.")
        return True
        
    except Exception as e:
        print(f"Ocorreu um erro ao inserir o valor da coluna B da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao inserir placa: {e}"
        workbook.save(file_path)
        return False

def click_placa_value(placa_value, row):
    try:
        print(f"Aguardando o dropdown menu com a lista de placas...")
        attempts = 3  
        
        while attempts > 0:
            try:
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, '//ul[@role="listbox"]'))
                )
                WebDriverWait(driver, 30).until(
                    EC.visibility_of_all_elements_located((By.XPATH, '//ul[@role="listbox"]/li'))
                )
                li_elements = driver.find_elements(By.XPATH, '//ul[@role="listbox"]/li')

                for li in li_elements:
                    a_element = li.find_element(By.XPATH, './/a')
                    strong_element = a_element.find_element(By.XPATH, './/strong')
                    strong_text = strong_element.text.strip()
                    full_text = a_element.text.strip()

                    if strong_text == placa_value:
                        if full_text == strong_text or (full_text.startswith(strong_text) and full_text[len(strong_text):].strip() == ''):
                            print(f"Elemento correspondente encontrado: '{full_text}'")
                            li.click()
                            print(f"Elemento com placa '{placa_value}' clicado com sucesso.")
                            return True 
                        
                print(f"Nenhum elemento correspondente encontrado com a placa '{placa_value}'.")

                if attempts > 1:
                    placa_field = driver.find_element(By.NAME, 'placa_vinculo')
                    if reset_search_input(placa_field, placa_value):
                        attempts -= 1
                    else:
                        sheet[f'G{row}'] = f"Erro ao tentar redefinir a busca para a placa: {placa_value}"
                        workbook.save(file_path)
                        return False
                else:
                    sheet[f'G{row}'] = f"Placa não encontrada: {placa_value}"
                    workbook.save(file_path)
                    return False

            except TimeoutException:
                print(f"Tempo esgotado ao esperar pelo dropdown com placas.")
                if attempts > 1:
                    placa_field = driver.find_element(By.NAME, 'placa_vinculo')
                    if reset_search_input(placa_field, placa_value):
                        attempts -= 1
                    else:
                        sheet[f'G{row}'] = f"Erro ao tentar redefinir a busca para a placa: {placa_value}"
                        workbook.save(file_path)
                        return False
                else:
                    sheet[f'G{row}'] = f"Placa não encontrada: {placa_value}"
                    workbook.save(file_path)
                    return False

    except Exception as e:
        print(f"Erro ao clicar na placa {placa_value} da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao clicar na placa: {e}"
        workbook.save(file_path)
        return False


def copy_and_insert_value():
    try:
        print("Verificando presença do elemento <span> específico com CPF...")
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//div[@class="col-sm-6" and @ng-show="formData.cliente.cpf"]//span[@class="form-control ng-binding"]'))
        )

        print("Copiando valor do <span> específico com CPF...")
        span_element = driver.find_element(By.XPATH, '//div[@class="col-sm-6" and @ng-show="formData.cliente.cpf"]//span[@class="form-control ng-binding"]')
        span_value = span_element.text  
        formatted_value = span_value.replace('.', '').replace('-', '').replace('/', '')
        print(f"Valor do CPF formatado: {formatted_value}")

        attempts = 3  
        while attempts > 0:
            try:
                print("Inserindo valor formatado do CPF no campo de texto...")
                input_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="formData.interveniente.nome"]')
                input_field.clear()
                input_field.send_keys(formatted_value)
                print("Valor do CPF inserido com sucesso no campo de texto.")
                
                print(f"Aguardando elemento <li> contendo o valor '{formatted_value}'...")
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, f"//li[contains(., '{formatted_value}')]"))
                )
                li_element = driver.find_element(By.XPATH, f"//li[contains(., '{formatted_value}')]")
                li_element.click()
                print(f"Elemento <li> com valor '{formatted_value}' clicado com sucesso.")
                return True  

            except TimeoutException:
                print("Tempo esgotado ao tentar inserir o valor ou clicar no elemento correspondente.")
                attempts -= 1
                if attempts > 0:
                    print("Tentando redefinir a entrada e tentar novamente...")
                    reset_search_input(input_field, formatted_value)  
                else:
                    print("Todas as tentativas falharam. Indo para o próximo valor.")
                    return False 

    except Exception as e:
        print("Ocorreu um erro ao copiar e inserir os valores:", e)
        driver.quit()
        exit()

def insert_serial_value(row):
    try:
        print(f"Inserindo valor da coluna D da linha {row}...")
        serial_value = sheet[f'D{row}'].value
        if not serial_value:
            print(f"Nenhum valor encontrado na coluna D da linha {row}. Encerrando.")
            driver.quit()
            exit()

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'rastreador_vinculo'))
        )
        serial_field = driver.find_element(By.NAME, 'rastreador_vinculo')
        serial_field.clear()
        serial_field.send_keys(serial_value)
        print(f"Valor '{serial_value}' inserido no campo de texto.")

    except Exception as e:
        print(f"Ocorreu um erro ao inserir o valor da coluna D da linha {row}:", e)
        driver.quit()
        exit()

def click_serial_value(serial_value, row):
    try:
        print(f"Aguardando o dropdown menu com a lista de seriais...")
        attempts = 3  
        while attempts > 0:
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, f"//li[contains(., '{serial_value}')]"))
                )
                
                serial_element = driver.find_element(By.XPATH, f"//li[contains(., '{serial_value}')]")
                if serial_element:  
                    serial_element.click()
                    print(f"Elemento com valor '{serial_value}' clicado com sucesso.")
                    return True  

                print(f"Nenhum elemento correspondente encontrado com o serial '{serial_value}'.")

                if attempts > 0:
                    serial_field = driver.find_element(By.NAME, 'rastreador_vinculo')
                    if reset_search_input(serial_field, serial_value):
                        attempts -= 1
                    else:
                        sheet[f'G{row}'] = f"Erro ao tentar redefinir a busca para o serial: {serial_value}"
                        workbook.save(file_path)
                        return False
                else:
                    sheet[f'G{row}'] = f"Serial não encontrado: {serial_value}"
                    workbook.save(file_path)
                    return False

            except TimeoutException:
                print(f"Tempo esgotado ao esperar pelo dropdown com seriais.")
                if attempts > 0:
                    serial_field = driver.find_element(By.NAME, 'rastreador_vinculo')
                    if reset_search_input(serial_field, serial_value):
                        attempts -= 1
                    else:
                        sheet[f'G{row}'] = f"Erro ao tentar redefinir a busca para o serial: {serial_value}"
                        workbook.save(file_path)
                        return False
                else:
                    sheet[f'G{row}'] = f"Serial não encontrado: {serial_value}"
                    workbook.save(file_path)
                    return False

    except Exception as e:
        print(f"Erro ao clicar no serial {serial_value} da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao clicar no serial: {e}"
        workbook.save(file_path)
        return False

def fill_installation_date():
    try:
        hoje = datetime.now().strftime("%d/%m/%Y")
        print(f"Inserindo a data de instalação: {hoje}")

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//input[@ng-model="formData.data_instalacao"]'))
        )
        date_field = driver.find_element(By.XPATH, '//input[@ng-model="formData.data_instalacao"]')
        date_field.clear()
        date_field.send_keys(hoje)
        print(f"Data de instalação '{hoje}' inserida com sucesso.")

    except Exception as e:
        print("Ocorreu um erro ao inserir a data de instalação:", e)
        driver.quit()
        exit()

def fill_installation_location(row):
    try:
        print(f"Inserindo valor da coluna E da linha {row} para Local de Instalação...")
        local_instalacao = sheet[f'E{row}'].value
        if not local_instalacao:
            print(f"Nenhum valor encontrado na coluna E da linha {row}. Pulando este campo.")
            return 

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//input[@ng-model="formData.local_instalacao_modulo"]'))
        )
        local_field = driver.find_element(By.XPATH, '//input[@ng-model="formData.local_instalacao_modulo"]')
        local_field.clear()
        local_field.send_keys(local_instalacao)
        print(f"Local de instalação '{local_instalacao}' inserido com sucesso.")
        
    except Exception as e:
        print(f"Ocorreu um erro ao inserir o valor da coluna E da linha {row}:", e)
        driver.quit()
        exit()

def select_technician(row):
    try:
        tecnico_nome = sheet[f'C{row}'].value.upper()
        if not tecnico_nome:
            print(f"Nenhum valor encontrado na coluna C da linha {row}. Encerrando.")
            driver.quit()
            exit()

        print(f"Buscando o técnico: {tecnico_nome}")
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//select[@ng-model="formData.tecnico_id"]'))
        )
        tecnico_select = driver.find_element(By.XPATH, '//select[@ng-model="formData.tecnico_id"]')
        
        tecnico_select.click()

        if tecnico_nome == "IAGO":
            print(f"Selecionando técnico específico: IAGO DANIEL LIMA OLIVEIRA...")
            options = driver.find_elements(By.XPATH, '//select[@ng-model="formData.tecnico_id"]/option')
            for option in options:
                if "IAGO DANIEL LIMA OLIVEIRA" in option.text:
                    option.click()
                    print(f"Técnico 'IAGO DANIEL LIMA OLIVEIRA' selecionado com sucesso.")
                    return  

        options = driver.find_elements(By.XPATH, '//select[@ng-model="formData.tecnico_id"]/option')
        
        tecnico_selecionado = False
        
        for option in options:
            if tecnico_nome in option.text:
                print(f"Tentando selecionar o técnico: {option.text}")
                option.click()

                if "REMOVIDOS" in option.text:
                    print(f"Técnico '{option.text}' contém 'REMOVIDOS', tentando o próximo...")
                    tecnico_select.click()  
                else:
                    tecnico_selecionado = True
                    print(f"Técnico '{option.text}' selecionado com sucesso.")
                    break 
                
        if not tecnico_selecionado:
            print(f"Nenhum técnico correspondente encontrado para '{tecnico_nome}'.")

    except Exception as e:
        print(f"Ocorreu um erro ao selecionar o técnico da coluna C da linha {row}:", e)
        driver.quit()
        exit()

def select_system():
    try:
        print("Selecionando o sistema 'GETRAK'...")
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//select[@ng-model="formData.sistema_id"]'))
        )

        system_select = driver.find_element(By.XPATH, '//select[@ng-model="formData.sistema_id"]')
        system_select.click() 
        
        option_getrak = driver.find_element(By.XPATH, '//select[@ng-model="formData.sistema_id"]/option[@value="string:270"]')
        option_getrak.click()
        print("Sistema 'GETRAK' selecionado com sucesso.")
        
    except Exception as e:
        print("Ocorreu um erro ao selecionar as opções:", e)
        driver.quit()
        exit()

def mark_checkbox_by_label_text(driver, label_text):
    try:
        label_text = label_text.strip()

        checkbox = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//label[contains(normalize-space(), '{label_text}')]/input[@type='checkbox']")
            )
        )

        driver.execute_script("arguments[0].scrollIntoView(true);", checkbox)

        if not checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)

            driver.execute_script("""
                var checkbox = arguments[0];
                var event = new Event('change', { bubbles: true });
                checkbox.dispatchEvent(event);
                
                var angularElement = angular.element(checkbox);
                var scope = angularElement.scope();
                var model = angularElement.attr('ng-model');
                scope.$apply(function() {
                    scope.$eval(model + ' = true');
                });
            """, checkbox)

            print(f"Checkbox com label '{label_text}' marcado com sucesso.")
        else:
            print(f"Checkbox com label '{label_text}' já estava marcado.")
    except Exception as e:
        print(f"Erro ao marcar a checkbox com label '{label_text}':", e)

def scroll_to_element(driver, element):
    try:
        driver.execute_script("window.scrollTo(0, arguments[0].getBoundingClientRect().top + window.pageYOffset - 100);", element)
    except Exception as e:
        print(f"Erro ao rolar para o elemento: {e}")        

def select_option_by_value(driver, select_locator, value):
    try:
        select_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, select_locator))
        )
        
        scroll_to_element(driver, select_element)
        WebDriverWait(driver, 5).until(
            EC.visibility_of(select_element)
        )
        
        if select_element.tag_name == "select":
            select = Select(select_element)
            select.select_by_value(value)
            print(f"Valor '{value}' selecionado no select com localizador '{select_locator}'.")
        else:
            print(f"Elemento com localizador '{select_locator}' não é um select.")
    except NoSuchElementException as e:
        print(f"Elemento não encontrado ao selecionar valor '{value}' no select com localizador '{select_locator}':", e)
    except TimeoutException as e:
        print(f"Tempo de espera expirado ao esperar pelo select com localizador '{select_locator}':", e)
    except Exception as e:
        print(f"Erro ao selecionar valor '{value}' no select com localizador '{select_locator}':", e)

def select_option_by_js(driver, select_locator, value):
    try:
        select_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, select_locator))
        )
        scroll_to_element(driver, select_element)
        WebDriverWait(driver, 5).until(
            EC.visibility_of(select_element)
        )
        
        if select_element.tag_name == "select":
            driver.execute_script(f"arguments[0].value='{value}';", select_element)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", select_element)

            driver.execute_script("""
                var select = arguments[0];
                var event = new Event('change', { bubbles: true });
                select.dispatchEvent(event);
                
                var angularElement = angular.element(select);
                var scope = angularElement.scope();
                var model = angularElement.attr('ng-model');
                scope.$apply(function() {
                    scope.$eval(model + ' = select.value');
                });
            """, select_element)

            print(f"Valor '{value}' selecionado no select com localizador '{select_locator}' usando JavaScript.")
        else:
            print(f"Elemento com localizador '{select_locator}' não é um select.")
    except Exception as e:
        print(f"Erro ao selecionar valor '{value}' no select com localizador '{select_locator}' usando JavaScript:", e)

def select_option_by_angular(driver, select_locator, value):
    try:
        select_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, select_locator))
        )
        
        driver.execute_script("arguments[0].scrollIntoView(true);", select_element)
        
        driver.execute_script("""
            var selectElement = arguments[0];
            var value = arguments[1];
            
            // Obter o elemento AngularJS
            var angularElement = angular.element(selectElement);
            
            // Obter o scope
            var scope = angularElement.scope();
            
            // Alterar o valor do modelo
            scope.$apply(function() {
                scope.formData.banco_id = value;
            });
            
            // Disparar o evento de mudança
            var event = new Event('change', { bubbles: true });
            selectElement.dispatchEvent(event);
        """, select_element, value)

        print(f"Valor '{value}' selecionado no select com localizador '{select_locator}' usando AngularJS.")
    except Exception as e:
        print(f"Erro ao selecionar valor '{value}' no select com localizador '{select_locator}' usando AngularJS:", e)

def click_gravar_button(driver):
    try:
        gravar_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@type='submit' and contains(@class, 'btn-success')]"))
        )
        
        driver.execute_script("arguments[0].scrollIntoView(true);", gravar_button)
        
        gravar_button.click()
        print("Botão 'Gravar' clicado com sucesso.")
    except Exception as e:
        print("Erro ao clicar no botão 'Gravar':", e)   

def verificar_resultado_sincronismo(driver, sheet, row, file_path):
    try:
        try:
            erro_vinculo = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@id='feedback']//ul/li[contains(text(), 'A Getrak não permite mais de um rastreador por veículo')]"))
            )
            sheet[f'G{row}'] = "Erro: Veículo já possui rastreador vinculado"
            print("Erro encontrado: Veículo já possui rastreador vinculado.")
        except TimeoutException:
            element = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located((By.XPATH, "//h3[@data-ng-bind-html='titulo' and contains(text(), 'Resultado Sincronismo')]"))
            )

            icone_warning = driver.find_elements(By.XPATH, "//i[contains(@class, 'fa-exclamation-triangle text-warning')]")
            icone_danger = driver.find_elements(By.XPATH, "//i[contains(@class, 'fa-ban text-danger')]")

            if icone_warning:
                sheet[f'G{row}'] = "Problema encontrado"
                print("Problema encontrado durante o sincronismo.")
            elif icone_danger:
                sheet[f'G{row}'] = "Erro: Rastreador vinculado a teste"
                print("Erro encontrado: Rastreador vinculado a teste.")
            else:
                sheet[f'G{row}'] = "Sincronismo realizado com sucesso"
                print("Sincronismo realizado com sucesso.")

            ok_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@class='modal-footer ng-scope']/button[contains(@class, 'btn-success') and text()='OK']"))
            )
            ok_button.click()

    except Exception as e:
        print(f"Erro ao verificar o resultado do sincronismo: {e}")
        sheet[f'G{row}'] = f"Erro: {str(e)}"

    finally:
        workbook.save(file_path)
        print("Arquivo Excel salvo com sucesso.")


def confirm_close_browser():

    print("Salvando o arquivo Excel...")
    workbook.save(file_path)
    print("Arquivo Excel salvo com sucesso.")

    response = input("Deseja fechar o navegador? (s/n): ").strip().lower()
    if response == 's':
        print("Salvando o arquivo Excel...")
        workbook.save(file_path)
        print("Arquivo Excel salvo com sucesso.")
        driver.quit()
        print("Navegador fechado.")
    else:
        print("Navegador não fechado. Você pode fechá-lo manualmente quando desejar.")

def main():
    try:
        open_page_and_accept()

        login()

        start_row = int(input("Digite a linha inicial da planilha: "))

        navigate_to_vincular_rastreador_page()

        for row in range(start_row, sheet.max_row + 1):
            insert_placa_value(row)
            placa_value = sheet[f'B{row}'].value
            print(f"Tentando clicar na placa: {placa_value}")
            click_result = click_placa_value(placa_value, row)

            if not click_result:  
                print("Erro ao entrar com a placa. Atualizando a página e passando para a próxima.")
                driver.refresh()  
                continue  

            copy_and_insert_value()
            insert_serial_value(row)
            serial_value = sheet[f'D{row}'].value

            if not click_serial_value(serial_value, row):
                print(f"Erro ao clicar no serial. Atualizando a página e passando para a próxima linha.")
                driver.refresh() 
                continue  
            
            fill_installation_date()
            fill_installation_location(row)
            select_technician(row)
            select_system()
            mark_checkbox_by_label_text(driver, 'Alimentação desconectada')
            
            select_option_by_value(driver, "//select[@ng-model='formData.telemetria_notificacao[subitem.option]']", "2_3")
            
            select_option_by_value(driver, "//select[@id='dia_vencimento']", "string:267")

            select_option_by_angular(driver, "//select[@id='banco']", "1")

            select_option_by_value(driver, "//select[@id='grupo']", "string:15")

            gravar_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-success') and text()='Gravar']"))
                )
            gravar_button.click()

            verificar_resultado_sincronismo(driver, sheet, row, file_path)
                
            driver.refresh()
                
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

            
    except Exception as e:
        print("Ocorreu um erro durante a execução do script:", e)
    finally:
        confirm_close_browser()

if __name__ == "__main__":
    main()
