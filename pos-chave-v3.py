from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from datetime import datetime
import openpyxl

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/POS CHAVE.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def wait_for_angular():
    """Aguardar o AngularJS concluir suas atualizações."""
    try:
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return window.angular !== undefined && angular.element(document).injector().get("$http").pendingRequests.length === 0')
        )
        print("AngularJS atualizado.")
    except Exception as e:
        print("Erro ao aguardar o AngularJS:", e)

def wait_until_element_clickable(by, value, timeout=10):
    """Espera até que o elemento especificado esteja clicável."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )

def open_page_and_accept():
    try:
        print("Abrindo a página de agendamento...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]'))
        )
        print("Página de agendamento carregada.")

        print("Clicando no botão 'Aceitar e fechar'...")
        accept_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]')
        accept_button.click()
        print("Botão 'Aceitar e fechar' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e clicar no aviso:", e)
        driver.quit()
        exit()

def login():
    try:
        print("Preenchendo o formulário de login...")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]'))
        )
        client_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]')
        client_field.clear()
        client_field.send_keys("")

        name_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.nome"]')
        name_field.clear()
        name_field.send_keys("")

        password_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.senha"]')
        password_field.clear()
        password_field.send_keys("")

        print("Clicando no botão 'Entrar'...")
        login_button = driver.find_element(By.CSS_SELECTOR, 'button[ng-click="login()"]')
        login_button.click()
        print("Botão 'Entrar' clicado com sucesso.")

        time.sleep(5)

    except Exception as e:
        print("Ocorreu um erro ao preencher o formulário de login e clicar em 'Entrar':", e)
        driver.quit()
        exit()

def navigate_to_veiculos_page():
    try:
        print("Navegando para a página de veiculos...")
        driver.get('site')
        print("Página de veiculos carregada.")
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="search"].form-control.input-sm'))
        )
        
    except Exception as e:
        print("Ocorreu um erro ao navegar para a página de vinculação de rastreador:", e)
        driver.quit()
        exit()

def insert_placa_value(row):
    try:
        print(f"Inserindo valor da coluna B da linha {row}...")
        placa_value = sheet[f'B{row}'].value
        if not placa_value:
            print(f"Nenhum valor encontrado na coluna B da linha {row}. Parando o processamento.")
            return False

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="search"].form-control.input-sm'))
        )

        placa_field = driver.find_element(By.CSS_SELECTOR, 'input[type="search"].form-control.input-sm')
        placa_field.clear()
        placa_field.send_keys(placa_value)
        placa_field.send_keys(Keys.RETURN) 

        print(f"Valor '{placa_value}' inserido no campo de texto e pesquisa iniciada.")

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.dataTable tbody tr'))
        )

        table_body = driver.find_element(By.CSS_SELECTOR, 'table.dataTable tbody')
        rows = table_body.find_elements(By.TAG_NAME, 'tr')

        placa_encontrada = False
        for tr in rows:
            cols = tr.find_elements(By.TAG_NAME, 'td')
            if cols and cols[0].text.strip() == placa_value.strip():
                placa_encontrada = True
                print(f"Placa '{placa_value}' encontrada na tabela.")
                
                button = tr.find_element(By.CSS_SELECTOR, 'button.btn-primary')
                button.click()

                WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.modal-content'))
                )

                time.sleep(2)

                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.modal-dialog .table-responsive tbody'))
                )

                modal_table_body = driver.find_element(By.CSS_SELECTOR, 'div.modal-dialog .table-responsive tbody')
                modal_rows = modal_table_body.find_elements(By.TAG_NAME, 'tr')

                if modal_rows:
                    tr_id = modal_rows[0].get_attribute('id')
                    if tr_id:
                        url_nova_pagina = f"site{tr_id}"
                        driver.get(url_nova_pagina)
                        print(f"Acessando a página: {url_nova_pagina}")

                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'form')) 
                        )

                        WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'select.form-control'))
                        )
                        
                        checkbox_status = unmark_checkbox_by_label_text(driver, "Alimentação desconectada")

                        print("Clicando no botão 'Gravar'...")
                        save_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success')
                        save_button.click()

                        WebDriverWait(driver, 20).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.modal-dialog'))
                        )

                        print("Clicando no botão 'OK' no modal de sucesso...")
                        ok_button = driver.find_element(By.CSS_SELECTOR, 'div.modal-dialog .btn.btn-success[ng-click="ok()"]')
                        ok_button.click()

                        WebDriverWait(driver, 10).until(
                            EC.invisibility_of_element_located((By.CSS_SELECTOR, 'div.modal-dialog'))
                        )

                        sheet[f'E{row}'] = checkbox_status
                        workbook.save(file_path)

                        navigate_to_veiculos_page()
                        return True
                    else:
                        print("ID do <tr> não encontrado.")
                        sheet[f'G{row}'] = "ID do <tr> não encontrado"
                        workbook.save(file_path)
                        navigate_to_veiculos_page()  
                        return True  
                        
        if not placa_encontrada:
            print(f"Placa '{placa_value}' não encontrada na tabela.")
            sheet[f'G{row}'] = "Placa não encontrada na tabela"
            workbook.save(file_path)
        return True  
        
    except Exception as e:
        print(f"Ocorreu um erro ao inserir o valor da coluna B da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao inserir placa: {e}"
        workbook.save(file_path)
        return True  

def unmark_checkbox_by_label_text(driver, label_text):
    try:
        label_text = label_text.strip()

        checkbox = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//label[contains(normalize-space(), '{label_text}')]/input[@type='checkbox']")
            )
        )

        driver.execute_script("arguments[0].scrollIntoView(true);", checkbox)

        if checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)

            driver.execute_script("""
                var checkbox = arguments[0];
                var event = new Event('change', { bubbles: true });
                checkbox.dispatchEvent(event);
                
                var angularElement = angular.element(checkbox);
                var scope = angularElement.scope();
                var model = angularElement.attr('ng-model');
                scope.$apply(function() {
                    scope.$eval(model + ' = false');
                });
            """, checkbox)

            print(f"Checkbox com label '{label_text}' desmarcado com sucesso.")
            return "Desmarcado"
        else:
            print(f"Checkbox com label '{label_text}' já estava desmarcado.")
            return "Já desmarcado"
    except Exception as e:
        print(f"Erro ao desmarcar a checkbox com label '{label_text}':", e)
        return "Erro ao desmarcar"

def main():
    try:
        open_page_and_accept()

        login()

        start_row = int(input("Digite a linha inicial da planilha: "))

        navigate_to_veiculos_page()

        for row in range(start_row, sheet.max_row + 1):
            print(f"Processando linha {row}...")
            if not insert_placa_value(row):
                print(f"Parando o processamento na linha {row}.")
                break
            navigate_to_veiculos_page()

    except Exception as e:
        print("Ocorreu um erro durante a execução do script:", e)
    finally:
        driver.quit()
        workbook.save(file_path)

if __name__ == "__main__":
    main()



