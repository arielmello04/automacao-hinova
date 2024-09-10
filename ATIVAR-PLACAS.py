from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException, NoSuchWindowException
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from datetime import datetime
import openpyxl

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/ATIVAR PLACAS.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def wait_for_angular():
    """Aguardar o AngularJS concluir suas atualizações."""
    try:
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return window.angular !== undefined && angular.element(document).injector().get("$http").pendingRequests.length === 0')
        )
        print("AngularJS atualizado.")
    except Exception as e:
        print("Erro ao aguardar o AngularJS:", e)

def wait_until_element_clickable(by, value, timeout=10):
    """Espera até que o elemento especificado esteja clicável."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )

def open_page_and_accept():
    try:
        print("Abrindo a página de ativação do veiculo...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]'))
        )
        print("Página de agendamento carregada.")

        print("Clicando no botão 'Aceitar e fechar'...")
        accept_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]')
        accept_button.click()
        print("Botão 'Aceitar e fechar' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e clicar no aviso:", e)
        driver.quit()
        exit()

def login():
    try:
        print("Preenchendo o formulário de login...")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]'))
        )
        client_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]')
        client_field.clear()
        client_field.send_keys("")

        name_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.nome"]')
        name_field.clear()
        name_field.send_keys("")

        password_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.senha"]')
        password_field.clear()
        password_field.send_keys("")

        print("Clicando no botão 'Entrar'...")
        login_button = driver.find_element(By.CSS_SELECTOR, 'button[ng-click="login()"]')
        login_button.click()
        print("Botão 'Entrar' clicado com sucesso.")

        time.sleep(5)

    except Exception as e:
        print("Ocorreu um erro ao preencher o formulário de login e clicar em 'Entrar':", e)
        driver.quit()
        exit()

def navigate_to_vincular_rastreador_page():
    try:
        print("Navegando para a página de vinculação de rastreador...")
        driver.get('site')
        print("Página de vinculação de rastreador carregada.")
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'placa_vinculo'))
        )
        
    except Exception as e:
        print("Ocorreu um erro ao navegar para a página de vinculação de rastreador:", e)
        driver.quit()
        exit()

def reset_search_input(input_field, value):
    try:
        input_field.send_keys(Keys.BACKSPACE) 
        input_field.send_keys(value[-1]) 
        print("Último caractere removido e reinserido.")
        return True
    except Exception as e:
        print(f"Erro ao tentar redefinir a entrada de busca: {e}")
        return False

def insert_plate_from_excel(row=2, column='A'):
    try:
        cell_value = sheet[f'{column}{row}'].value

        if cell_value is None:
            print(f"Não há valor na célula {column}{row}.")
            return None

        plate_value = str(cell_value).strip() 
        print(f"Inserindo a placa '{plate_value}' no campo de pesquisa...")

        search_field = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="search"]'))
        )

        print("Campo de pesquisa encontrado. Inserindo a placa...")

        search_field.clear()
        search_field.send_keys(plate_value)
        print(f"Placa '{plate_value}' inserida com sucesso.")

        time.sleep(3) 
        return plate_value

    except TimeoutException:
        print(f"Erro: Tempo esgotado ao tentar encontrar o campo de pesquisa para a placa {cell_value}.")
        sheet[f'C{row}'] = "Erro de timeout ao inserir placa"
        workbook.save(file_path)
        return None
    except Exception as e:
        print(f"Ocorreu um erro ao inserir a placa: {e}")
        sheet[f'C{row}'] = f"Erro ao inserir a placa: {str(e)}"
        workbook.save(file_path)
        return None

def wait_for_table_and_click_plate(plate_value, row):
    try:
        print(f"Aguardando a tabela carregar com a placa {plate_value}...")
        
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'tbody tr'))
        )

        rows = driver.find_elements(By.CSS_SELECTOR, 'tbody tr')
        for row_elem in rows:
            label = row_elem.find_element(By.CSS_SELECTOR, 'td label').text.strip()

            if label == plate_value:
                print(f"Placa exata '{plate_value}' encontrada na tabela.")
                
                status_span = row_elem.find_element(By.CSS_SELECTOR, 'span[title="Alterar situação"] span.label')
                current_status = status_span.text.strip()
                print(f"Status atual da placa {plate_value}: {current_status}")

                if current_status != "ATIVO":
                    status_span.click()
                    print(f"Alterando a situação da placa {plate_value}...")

                    change_status_to_active()
                    
                    sheet[f'C{row}'] = "Ativado com sucesso"
                    workbook.save(file_path)
                    return True
                else:
                    print(f"Placa {plate_value} já está ATIVA.")
                    sheet[f'C{row}'] = "Já está ATIVO"
                    workbook.save(file_path)
                    return False

        print(f"Placa exata '{plate_value}' não encontrada na tabela.")
        sheet[f'C{row}'] = "Placa não encontrada"
        workbook.save(file_path)
        return False

    except TimeoutException:
        print(f"Erro: Tempo esgotado aguardando a tabela para a placa {plate_value}.")
        sheet[f'C{row}'] = "Erro de timeout"
        workbook.save(file_path)
        return False
    except Exception as e:
        print(f"Erro ao procurar a placa na tabela: {e}")
        sheet[f'C{row}'] = f"Erro: {str(e)}"
        workbook.save(file_path)
        driver.quit()
        exit()

def change_status_to_active():
    try:
        print("Aguardando o modal de alteração de situação abrir...")
        WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.modal-content'))
        )
        
        situation_select_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'select[ng-model="formData.situacao_id"]'))
        )
        
        situation_select = Select(situation_select_element)
        options = [option.text for option in situation_select.options]

        retries = 0
        while len(options) <= 1 and retries < 5:  
            print("Aguardando o carregamento das opções do dropdown...")
            time.sleep(2)  
            situation_select = Select(situation_select_element)
            options = [option.text for option in situation_select.options]
            retries += 1
        
        print(f"Opções encontradas no dropdown: {options}")

        if "ATIVO" in options:
            print("Selecionando 'ATIVO' como nova situação...")
            situation_select.select_by_visible_text('ATIVO')
        else:
            print("Erro: a opção 'ATIVO' não está disponível no dropdown.")
            return

        print("Clicando no botão 'Gravar'...")
        gravar_button = driver.find_element(By.CSS_SELECTOR, 'button[type="submit"].btn-success')
        gravar_button.click()

        print("Aguardando confirmação de sincronismo...")
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.modal-footer button.btn-success'))
            )
            ok_button = driver.find_element(By.CSS_SELECTOR, 'div.modal-footer button.btn-success')
            ok_button.click()
            print("Sincronismo confirmado e fechado com sucesso.")
        except TimeoutException:
            print("Atenção: O modal de sincronismo não apareceu. Prosseguindo para a próxima placa.")

    except Exception as e:
        print(f"Erro ao alterar a situação do veículo: {e}")
        driver.quit()
        exit()

def process_plate(plate_value, row):
    if plate_value is None:
        print("Nenhuma placa fornecida.")
        return
    
    if wait_for_table_and_click_plate(plate_value, row):
        print("Atualizando a página...")
        driver.refresh()
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, 'body'))
        )
        print("Página atualizada.")
    else:
        print(f"Placa {plate_value} não foi processada com sucesso.")

def confirm_close_browser():
    try:
        print("Salvando o arquivo Excel...")
        workbook.save(file_path)
        print("Arquivo Excel salvo com sucesso.")

        response = input("Deseja fechar o navegador? (s/n): ").strip().lower()
        if response == 's':
            print("Fechando o navegador...")
            driver.quit()
            print("Navegador fechado.")
        else:
            print("Navegador não fechado. Você pode fechá-lo manualmente quando desejar.")
    except ValueError:
        print("Erro: Operação de I/O em arquivo fechado.")

def main():
    try:
        open_page_and_accept()

        login()

        print("Navegando para a página de cadastro de veículos...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, 'body'))
        )
        print("Página de cadastro de veículos carregada com sucesso.")

        start_row = 2 
        end_row = sheet.max_row  

        for row in range(start_row, end_row + 1):
            plate_value = insert_plate_from_excel(row)

            if plate_value is None:
                print(f"Erro ao inserir a placa da linha {row}. Pulando para a próxima...")
                continue

            process_plate(plate_value, row)

            print("Atualizando a página...")
            driver.refresh()
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
            )
            print("Página atualizada para a próxima placa.")

    except Exception as e:
        print(f"Ocorreu um erro durante a execução do script: {e}")
    finally:
        try:
            confirm_close_browser()
        except Exception as e:
            print(f"Erro durante o fechamento do navegador e salvamento do arquivo: {e}")

if __name__ == "__main__":
    main()
