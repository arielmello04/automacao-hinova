from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
import time

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/PREENCHER DADOS.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def save_page_html(filename):
    with open(filename, "w", encoding="utf-8") as file:
        file.write(driver.page_source)
    print(f"HTML da página salvo como {filename}")

def login():
    try:
        print("Abrindo o site Saturno Hinova...")
        driver.get('site')

        try:
            print("Aguardando o botão 'Continuar e Fechar' aparecer...")
            accept_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.btn.btn-primary.aceite-cookie'))
            )
            accept_button.click()
            print("Botão 'Continuar e Fechar' clicado com sucesso.")
        except Exception as e:
            print("O botão 'Continuar e Fechar' não apareceu ou não pôde ser clicado:", e)
        
        print("Preenchendo o formulário de login...")
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'usuario'))
        )
        password_field = driver.find_element(By.ID, 'senha')
        login_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-primary')

        username_field.send_keys('')
        password_field.send_keys('')
        login_button.click()
        print("Formulário de login preenchido e enviado.")
        
        print("Aguardando o botão 'Fechar' aparecer...")
        close_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.btn.btn-default[data-dismiss="modal"]'))
        )
        time.sleep(1)
        close_button.click()
        print("Botão 'Fechar' clicado com sucesso.")
        
        time.sleep(2)
        print("Abrindo a página de pesquisa de veículos...")
        driver.get('site')

    except Exception as e:
        print("Ocorreu um erro durante o login:", e)
        save_page_html("login_error_page.html")
        driver.quit()
        exit()

def close_popup_if_present():
    try:
        overlay = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.fancybox-overlay'))
        )
        
        if overlay.is_displayed():
            print("Pop-up 'Atenção' detectado. Fechando o pop-up...")
            close_button = driver.find_element(By.CSS_SELECTOR, 'a.fancybox-close')
            close_button.click()
            print("Pop-up fechado com sucesso.")
    
    except Exception as e:
        print("Pop-up 'Atenção' não foi encontrado ou não pôde ser fechado")
    
    while True:
        try:
            alert = WebDriverWait(driver, 2).until(EC.alert_is_present())
            if alert:
                print("Alerta detectado. Fechando o alerta...")
                alert_text = alert.text
                print(f"Texto do alerta: {alert_text}")
                alert.accept()
                print("Alerta fechado com sucesso.")
        except Exception as e:
            print("Nenhum alerta foi encontrado ou houve um erro ao fechá-lo")
            break

def process_plate(plate, row_number):
    retry_count = 0
    max_retries = 5 

    while retry_count < max_retries:
        try:
            if len(plate) > 7:
                search_field_id = 'dfsChassiFiltro'
                autosuggest_id = 'as_dfsChassiFiltro'
            else:
                search_field_id = 'dfsPlacaFiltro'
                autosuggest_id = 'as_dfsPlacaFiltro'

            print(f"Processando a placa/chassi: {plate}")
            driver.get('site')

            print(f"Aguardando o campo de pesquisa '{search_field_id}'...")
            search_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, search_field_id))
            )
            search_field.clear()
            search_field.send_keys(plate)
            print(f"Campo de pesquisa preenchido com '{plate}'.")

            print(f"Aguardando a lista de sugestões '{autosuggest_id}' aparecer...")
            suggestions_list = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, f'#{autosuggest_id}'))
            )
            suggestions = suggestions_list.find_elements(By.TAG_NAME, 'a')
            if suggestions:
                print(f"Selecionando o primeiro item da lista de sugestões: {suggestions[0].text}")
                suggestions[0].click()
                print("Primeiro item da lista de sugestões clicado.")
            else:
                print("Nenhuma sugestão encontrada.")
                sheet[f'D{row_number}'].value = 'Sugestão não encontrada'
                workbook.save(file_path)
                return

            close_popup_if_present()

            print("Extraindo informações do proprietário...")
            try:
                owner_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@width='30%' and @align='left' and @class='label']"))
                )
                owner_name = owner_element.text.strip()
                print(f"Nome do proprietário: {owner_name}")
            except Exception as e:
                print("Não foi possível extrair o nome do proprietário:", e)
                owner_name = 'Não encontrado'

            print("Extraindo informações da matrícula...")
            try:
                matricula_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@width='8%' and @align='left' and @class='label']//div[@align='left' and @class='label']"))
                )
                matricula = matricula_element.text.strip()
                print(f"Matrícula: {matricula}")
            except Exception as e:
                print("Não foi possível extrair a matrícula:", e)
                matricula = 'Não encontrado'

            print("Extraindo informações da cidade de correspondência...")
            try:
                city_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'dfsCidadeCorrespondencia'))
                )
                city = city_element.get_attribute('value')
                print(f"Cidade de correspondência: {city}")
            except Exception as e:
                print("Não foi possível extrair a cidade de correspondência:", e)
                city = 'Não encontrado'

            print(f"Escrevendo os valores na planilha para a placa: {plate}")
            sheet[f'A{row_number}'].value = matricula
            sheet[f'B{row_number}'].value = owner_name
            sheet[f'D{row_number}'].value = city  
            
            print("Alterando o valor do select 'cmbTipoAdesao' para 'COM / RASTREADOR'...")
            adesao_select = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'cmbTipoAdesao'))
            )
            adesao_option = adesao_select.find_element(By.CSS_SELECTOR, 'option[value="1"]')
            adesao_option.click()
            print("Valor do select alterado com sucesso.")

            dfs_data_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'dfsData'))
            )
            dfs_data_field.send_keys(Keys.ENTER)
            print("Seleção confirmada pressionando Enter.")
            
            while True:
                try:
                    alert = WebDriverWait(driver, 10).until(EC.alert_is_present())
                    alert_text = alert.text
                    print(f"Texto do alerta: {alert_text}")
                    alert.accept()
                    print("Alerta aceito com sucesso.")
                    
                    if "You have an error in your SQL syntax" in alert_text:
                        print("Alerta de erro SQL detectado. Atualizando a página e tentando novamente...")
                        driver.refresh() 
                        retry_count += 1 
                        break 
                    else:
                        continue
                except Exception as e:
                    print("Nenhum alerta de erro SQL foi encontrado. Procedendo...")
                    sheet[f'E{row_number}'].value = 'Alterado com sucesso'
                    workbook.save(file_path)
                    return 

            close_popup_if_present()
        except Exception as e:
            print(f"Ocorreu um erro ao processar a placa/chassi {plate}: {e}")
            sheet[f'E{row_number}'].value = 'Erro geral'
            workbook.save(file_path)
            break

    if retry_count == max_retries:
        print(f"Máximo de tentativas atingido para a placa/chassi {plate}.")
        sheet[f'E{row_number}'].value = 'Erro SQL após múltiplas tentativas'
        workbook.save(file_path)

def main(start_row):
    login()

    try:
        for row_number in range(start_row, sheet.max_row + 1):
            plate = sheet[f'C{row_number}'].value
            if not plate:
                print("Nenhuma placa encontrada na linha", row_number, ". Encerrando o processamento.")
                break
            process_plate(plate, row_number)
            time.sleep(2)

        workbook.save(file_path)
        print("Arquivo Excel salvo com sucesso.")
    except Exception as e:
        print("Ocorreu um erro durante o processamento:", e)
    finally:
        input("Pressione Enter para encerrar a execução e fechar o navegador...")
        driver.quit()

try:
    start_row = int(input("Digite a linha inicial para começar o processamento das placas: "))
except ValueError:
    print("Valor inválido para a linha inicial. Encerrando.")
    exit()

main(start_row)
