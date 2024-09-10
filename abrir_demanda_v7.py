from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains

driver_path = 'C:/Users/amelo/Downloads/chromedriver-win64/chromedriver.exe'

file_path = 'C:/Users/amelo/Desktop/ABRIR DEMANDA.xlsx'

print("Carregando o arquivo Excel...")
workbook = load_workbook(filename=file_path, read_only=False)
sheet = workbook.active

service = Service(driver_path)
driver = webdriver.Chrome(service=service)

def wait_until_element_clickable(by, value, timeout=10):
    """Espera até que o elemento especificado esteja clicável."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )

def open_page_and_accept():
    try:       
        print("Abrindo a página de agendamento...")
        driver.get('site')

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]'))
        )
        print("Página de agendamento carregada.")

        print("Clicando no botão 'Aceitar e fechar'...")
        accept_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-success[ng-click="ok()"]')
        accept_button.click()
        print("Botão 'Aceitar e fechar' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página e clicar no aviso:", e)
        sheet['D1'] = f"Erro ao abrir a página: {e}" 
        workbook.save(file_path)
        driver.quit()
        exit()

def login():
    try:
        print("Preenchendo o formulário de login...")
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]'))
        )
        client_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.cliente"]')
        client_field.clear()
        client_field.send_keys("")

        name_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.nome"]')
        name_field.clear()
        name_field.send_keys("")
        
        password_field = driver.find_element(By.CSS_SELECTOR, 'input[ng-model="usuario.senha"]')
        password_field.clear()
        password_field.send_keys("")
        
        print("Clicando no botão 'Entrar'...")
        login_button = driver.find_element(By.CSS_SELECTOR, 'button[ng-click="login()"]')
        login_button.click()
        print("Botão 'Entrar' clicado com sucesso.")

        time.sleep(3)

    except Exception as e:
        print("Ocorreu um erro ao preencher o formulário de login e clicar em 'Entrar':", e)
        sheet['D1'] = f"Erro ao fazer login: {e}"  
        workbook.save(file_path)
        driver.quit()
        exit()

def open_agendamento_page():
    try:
        print("Navegando para a página de agendamento...")
        driver.get('site')

        print("Aguardando o botão 'Cadastrar novo' ficar disponível...")
        new_registration_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[ui-sref="app.form.agendamento"]'))
        )
        print("Botão 'Cadastrar novo' encontrado. Clicando...")
        new_registration_button.click()
        print("Botão 'Cadastrar novo' clicado com sucesso.")

        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao abrir a página de agendamento e clicar no botão:", e)
        sheet['D1'] = f"Erro ao abrir a página de agendamento: {e}"  
        workbook.save(file_path)
        driver.quit()
        exit()        

def select_service_instalacao():
    try:
        print("Aguardando o campo de seleção de serviço ficar disponível...")
        service_select = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "servico"))
        )
        print("Campo de seleção de serviço encontrado.")

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "servico"))
        )

        # Aguardar até que a opção 'INSTALAÇÃO' esteja presente
        print("Aguardando a opção 'INSTALAÇÃO' estar disponível...")
        install_option = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//select[@id='servico']/option[@value='string:262']"))
        )
        print("Opção 'INSTALAÇÃO' disponível.")

        print("Selecionando a opção 'INSTALAÇÃO'...")
        select = Select(service_select)
        select.select_by_visible_text('INSTALAÇÃO')
        print("Opção 'INSTALAÇÃO' selecionada com sucesso.")

        selected_option = select.first_selected_option.text
        if selected_option == 'INSTALAÇÃO':
            print("Confirmação: A opção 'INSTALAÇÃO' está selecionada.")
        else:
            print(f"A opção selecionada é: {selected_option}. A opção 'INSTALAÇÃO' não foi selecionada.")
        
        time.sleep(2)

    except Exception as e:
        print("Ocorreu um erro ao selecionar a opção 'INSTALAÇÃO':", e)
        sheet['D1'] = f"Erro ao selecionar a opção 'INSTALAÇÃO': {e}"  
        workbook.save(file_path)
        driver.quit()
        exit()

def fill_plate(row):
    try:
        print(f"Inserindo placa da linha {row}...")
        placa_value = sheet[f'B{row}'].value
        if not placa_value:
            print(f"Nenhum valor encontrado na coluna B da linha {row}. Continuando para a próxima linha.")
            sheet[f'G{row}'] = "Nenhum valor encontrado"
            workbook.save(file_path)
            return False
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'placa'))
        )
        placa_field = driver.find_element(By.NAME, 'placa')
        placa_field.clear()
        placa_field.send_keys(placa_value)
        print(f"Valor '{placa_value}' inserido no campo de texto.")

        if click_plate_value(placa_value, row):
            print(f"Placa '{placa_value}' selecionada com sucesso.")
            return True
        else:
            print(f"Falha ao selecionar a placa '{placa_value}'.")
            return False

    except Exception as e:
        print(f"Ocorreu um erro ao preencher e selecionar a placa da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao preencher e selecionar placa: {e}"
        workbook.save(file_path)
        return False

def click_plate_value(placa_value, row):
    try:
        print(f"Aguardando a lista de placas...")
        attempts = 3 
        while attempts > 0:
            try:
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, '//ul[@role="listbox"]'))
                )
                WebDriverWait(driver, 30).until(
                    EC.visibility_of_all_elements_located((By.XPATH, '//ul[@role="listbox"]/li'))
                )
                li_elements = driver.find_elements(By.XPATH, '//ul[@role="listbox"]/li')

                for index, li in enumerate(li_elements):
                    a_element = li.find_element(By.XPATH, './/a')
                    strong_element = a_element.find_element(By.XPATH, './/strong')
                    strong_text = strong_element.text.strip()
                    full_text = a_element.text.strip()

                    print(f"Elemento <li> {index}: '{full_text}'")
                    print(f"Texto dentro <strong>: '{strong_text}'")
                    print(f"Texto fora do <strong>: '{full_text.replace(strong_text, '').strip()}'")

                for li in li_elements:
                    a_element = li.find_element(By.XPATH, './/a')
                    strong_element = a_element.find_element(By.XPATH, './/strong')
                    strong_text = strong_element.text.strip()
                    full_text = a_element.text.strip()

                    if strong_text == placa_value:
                        if full_text == strong_text or (full_text.startswith(strong_text) and full_text[len(strong_text):].strip() == ''):
                            print(f"Elemento correspondente encontrado: '{full_text}'")
                            a_element.click()
                            print(f"Elemento com placa '{placa_value}' clicado com sucesso.")
                            return True 

                print(f"Nenhum elemento correspondente encontrado com a placa '{placa_value}'.")

                if attempts > 0:
                    placa_field = driver.find_element(By.NAME, 'placa')
                    placa_field.clear()
                    placa_field.send_keys(placa_value[:-1])  
                    time.sleep(1)
                    placa_field.send_keys(placa_value[-1])  
                    time.sleep(1)
                    attempts -= 1
                else:
                    sheet[f'G{row}'] = f"Placa não encontrada: {placa_value}"
                    workbook.save(file_path)
                    return False

            except TimeoutException:
                print(f"Tempo esgotado ao esperar pela lista de placas.")
                if attempts > 0:
                    placa_field = driver.find_element(By.NAME, 'placa')
                    placa_field.clear()
                    placa_field.send_keys(placa_value[:-1])  
                    time.sleep(1)
                    placa_field.send_keys(placa_value[-1])  
                    time.sleep(1)
                    attempts -= 1
                else:
                    sheet[f'G{row}'] = f"Placa não encontrada: {placa_value}"
                    workbook.save(file_path)
                    return False

    except Exception as e:
        print(f"Erro ao clicar na placa {placa_value} da linha {row}:", e)
        sheet[f'G{row}'] = f"Erro ao clicar na placa: {e}"
        workbook.save(file_path)
        return False

def fill_date_and_time():
    try:
        print("Preenchendo os campos de data e hora...")
        today_date = datetime.now()
        today_date_str = today_date.strftime("%d/%m/%Y")
        future_date_str = (today_date + timedelta(days=15)).strftime("%d/%m/%Y")

        data_inicial_field = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, 'data_inicial'))
        )
        data_inicial_field.clear()
        data_inicial_field.send_keys(today_date_str)
        print(f"Data Inicial preenchida com '{today_date_str}'.")

        data_final_field = driver.find_element(By.NAME, 'data_final')
        data_final_field.clear()
        data_final_field.send_keys(future_date_str)
        print(f"Data Final preenchida com '{future_date_str}'.")

        horario_inicial_field = driver.find_element(By.NAME, 'horario_inicial')
        horario_inicial_field.clear()
        horario_inicial_field.send_keys('000000')
        print("Horário Inicial preenchido com '000000'.")

        horario_final_field = driver.find_element(By.NAME, 'horario_final')
        horario_final_field.clear()
        horario_final_field.send_keys('000000')
        print("Horário Final preenchido com '000000'.")

        time.sleep(2)
    except Exception as e:
        print(f"Ocorreu um erro ao preencher os campos de data e hora: {e}")
        sheet['D1'] = f"Erro ao preencher data e hora: {e}"
        workbook.save(file_path)

def select_technician(row):
    time.sleep(2)
    max_attempts = 3 
    tecnico_nome = sheet[f'C{row}'].value.upper()

    if not tecnico_nome:
        print(f"Nenhum valor encontrado na coluna C da linha {row}. Encerrando.")
        sheet[f'F{row}'] = "Nenhum técnico encontrado"
        workbook.save(file_path)
        return False

    for attempt in range(max_attempts):
        try:
            time.sleep(1)
            print(f"Tentativa {attempt + 1}: Buscando o técnico: {tecnico_nome}")

            tecnico_select = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//select[@ng-model="formData.tecnico_id"]'))
            )
            
            tecnico_select.click()  
            print("Dropdown do técnico aberto.")

            tecnico_selecionado = False
            tecnico_valor = None

            if tecnico_nome == "IAGO":
                tecnico_valor = 'string:67'
                print(f"Selecionando técnico específico: IAGO DANIEL LIMA OLIVEIRA...")
                driver.execute_script("""
                    var selectElement = arguments[0];
                    selectElement.value = arguments[1];
                    angular.element(selectElement).triggerHandler('change');
                    angular.element(document.body).injector().get('$rootScope').$apply();
                """, tecnico_select, tecnico_valor)
                tecnico_selecionado = True
            else:
                options = driver.find_elements(By.XPATH, '//select[@ng-model="formData.tecnico_id"]/option')
                for option in options:
                    if tecnico_nome in option.text.upper():
                        tecnico_valor = option.get_attribute('value')
                        print(f"Tentando selecionar o técnico: {option.text}")
                        driver.execute_script("""
                            var selectElement = arguments[0];
                            selectElement.value = arguments[1];
                            angular.element(selectElement).triggerHandler('change');
                            angular.element(document.body).injector().get('$rootScope').$apply();
                        """, tecnico_select, tecnico_valor)
                        tecnico_selecionado = True
                        print(f"Técnico '{option.text}' selecionado com sucesso.")
                        break 

            if tecnico_selecionado:
                selected_value = tecnico_select.get_attribute('value')
                if selected_value == tecnico_valor:
                    print(f"Técnico '{tecnico_nome}' definitivamente selecionado.")
                    sheet[f'F{row}'] = "Selecionado"
                    workbook.save(file_path)
                    return True 
                else:
                    print(f"Falha ao selecionar o técnico '{tecnico_nome}'. Verificação final falhou.")
            else:
                print(f"Técnico '{tecnico_nome}' não encontrado ou não selecionado.")
            
            if attempt + 1 == max_attempts:
                raise Exception(f"Falha ao selecionar o técnico '{tecnico_nome}' após {max_attempts} tentativas.")

        except (StaleElementReferenceException, TimeoutException, NoSuchElementException) as e:
            print(f"Erro ao tentar selecionar o técnico na tentativa {attempt + 1}: {e}")
            if attempt + 1 == max_attempts:
                print(f"Erro persistente ao tentar selecionar o técnico '{tecnico_nome}'. Encerrando.")
                sheet[f'F{row}'] = f"Erro ao selecionar técnico: {e}"
                workbook.save(file_path)
                return False

    return False 
        
def get_starting_row():
    while True:
        try:
            start_row = int(input("Qual linha da planilha você deseja começar? "))
            if start_row <= 0:
                print("Por favor, insira um número válido maior que 0.")
            else:
                return start_row
        except ValueError:
            print("Entrada inválida. Por favor, insira um número.")

def confirm_and_proceed(row):
    try:
        if sheet[f'F{row}'].value == "Selecionado":
            print("Clicando no botão 'Gravar'...")
            submit_button = driver.find_element(By.CSS_SELECTOR, 'button[type="submit"].btn-success')
            submit_button.click()

            print("Aguardando a modal de confirmação...")
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.modal-dialog'))
            )
            sheet[f'D{row}'] = "Demanda aberta com sucesso."
            workbook.save(file_path)

            print("Clicando no botão 'OK' na modal...")
            ok_button = driver.find_element(By.CSS_SELECTOR, 'div.modal-footer .btn-success')
            ok_button.click()

            print("Salvando o arquivo Excel...")
            workbook.save(file_path)
            print("Arquivo Excel salvo com sucesso.")

            time.sleep(3)

            print("Atualizando a página para nova operação...")
            driver.refresh()

            print("Navegando para a página de cadastro de agendamento...")
            driver.get('site')

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "servico"))
            )
            print("Página de cadastro de agendamento carregada com sucesso, pronta para nova operação.")
        else:
            print(f"Não foi possível prosseguir porque o técnico '{sheet[f'C{row}'].value}' não foi selecionado.")
    except Exception as e:
        print(f"Ocorreu um erro ao confirmar e prosseguir: {e}")
        sheet['D1'] = f"Erro ao confirmar e prosseguir: {e}"
        workbook.save(file_path)
        driver.quit()
        exit()

def main():
    try:
        open_page_and_accept()
        login()
        open_agendamento_page()

        start_row = get_starting_row()

        for row in range(start_row, sheet.max_row + 1):
            plate = sheet[f'B{row}'].value 
            technician_name = sheet[f'C{row}'].value 
            
            if plate: 
                print(f"Processando a placa: {plate}")
                select_service_instalacao()    
                if fill_plate(row):
                    
                    fill_date_and_time()
                    
                    if technician_name:
                        print(f"Tentando selecionar o técnico '{technician_name}'...")
                        select_technician(row)  
                        confirm_and_proceed(row)  
                    else:
                        print(f"Nome do técnico não encontrado na linha {row}.")
                else:
                    print(f"Falha ao processar a placa '{plate}'.")
            else:
                print(f"Placa na linha {row} está vazia. Pulando para a próxima linha.")
    
    except Exception as e:
        print(f"Ocorreu um erro durante a execução principal: {e}")
        sheet['D1'] = f"Erro durante a execução principal: {e}" 
    
    finally:
        print("Salvando o arquivo Excel...")
        workbook.save(file_path)
        print("Arquivo Excel salvo com sucesso.")

        input("Pressione Enter para encerrar a execução e fechar o navegador...")
        driver.quit()

if __name__ == "__main__":
    main()
